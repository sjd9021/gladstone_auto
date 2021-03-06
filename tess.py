import cv2
import pytesseract
import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment


class Tesseract():

    def __init__(self, file):
        self.file = file
        img = cv2.imread(self.file)
        self.text1 = pytesseract.image_to_string(img)
        self.text_lines = self.text1.replace(" ", "")
        self.text_lines = self.text_lines.lower()
        self.text_lines = self.text_lines.splitlines()
        while "" in self.text_lines:
            self.text_lines.remove("")

    def date(self):
        survey_date = self.text_lines[5]
        if survey_date == 'invoice' or 'cin' in survey_date:
            survey_date = self.text_lines[6]
        return survey_date

    def insurers(self):
        for i in self.text_lines:
            if 'ina/cwith:' in i:
                return i[10:].upper()

    def policy_num(self):
        for i in self.text_lines:
            if 'policyno.:' in i:
                return i[10:].upper()

    def total_amount(self):
        a = self.text1.rfind('$')
        return self.text1[a:a + 7]

    def branch(self):
        lists = []

        with open('sam1.csv', 'r') as file:
            x = file.read()
            x = os.linesep.join([s for s in x.splitlines() if s])
            x = x.splitlines()

            for i in x:
                if not i.isnumeric():
                    lists.append(i)
            insurer_branch = set(lists)

        for i in insurer_branch:
            if i.lower() in self.text1.lower():
                return i.upper()

    def our_ref(self):
        for i in self.text_lines:
            if 'ourref.:' in i:
                return i[8:].upper()
            if 'ourref:' in i:
                return i[7:].upper()
        return None


directory = 'folder'
for entry in os.scandir(directory):
    call = Tesseract(entry.path)
    our_ref = call.our_ref()
    branch = call.branch()
    insurers = call.insurers()
    date = call.date()
    total_amount = call.total_amount()
    policy = call.policy_num()
    print(our_ref)
    counter = 0
    df = pd.read_excel(
        'Remittance Tally Master for PB 12.07.2020 2.xlsx',
        sheet_name='Open')
    df2 = pd.read_excel(
        'Remittance Tally Master for PB 12.07.2020 2.xlsx',
        sheet_name='Closed')
    a = df.iloc[:, 1]
    b = df.iloc[:, 2]
    a2 = df2.iloc[:, 1]
    b2 = df2.iloc[:, 2]
    serial = df['SL NO'].iloc[-1] + 1
    b2 = b2.str.replace(" ", "")
    a2 = a2.str.replace(" ", "")
    a = a.str.replace(" ", "")
    b = b.str.replace(" ", "")

    for i in a:
        if our_ref == i:
            counter = 1
            print(our_ref + " This file exists in Open sheet" + i)

    for i in b:
        if our_ref == i:
            counter = 1
            print(our_ref + " This file exists in the Open sheet" + i)

    for i in a2:
        if our_ref == i:
            counter = 1
            print(our_ref + 'This file exists in closed sheet')

    for i in b2:
        if our_ref == i:
            counter = 1
            print(our_ref + 'This file exists in closed sheet')

    if our_ref == 'None':
        counter = 1
        print(f"error in  {our_ref}  please input manually ")

    if counter == 0:
        wb = openpyxl.load_workbook(
            'Remittance Tally Master for PB 12.07.2020 2.xlsx')
        ws = wb['Open']
        newRowLocation = ws.max_row + 1
        print(newRowLocation)
        ws.cell(column=1, row=newRowLocation, value=serial)
        x = 1
        for i in our_ref:
            if i == 'G' or i == 'N':
                ws.cell(column=2, row=newRowLocation, value=our_ref)
                x = 0
        if x == 1:
            ws.cell(column=3, row=newRowLocation, value=our_ref)

        ws.cell(column=4, row=newRowLocation, value=insurers)
        ws.cell(column=5, row=newRowLocation, value=branch)
        ws.cell(column=6, row=newRowLocation, value=policy)
        ws.cell(column=7, row=newRowLocation, value=date)
        ws.cell(column=9, row=newRowLocation, value=total_amount)
        ws.cell(column=10, row=newRowLocation, value=total_amount)
        ws.cell(column=13, row=newRowLocation, value='OPEN')
        ft1 = Font(name='Arial', size=12)

        for colNum in range(1, ws.max_column + 1):
            ws.cell(row=newRowLocation, column=colNum).font = ft1
            ws.cell(
                row=newRowLocation,
                column=colNum).alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
        wb.save('Remittance Tally Master for PB 12.07.2020 2.xlsx')
        wb.close()
